import openpyxl

from Config.settings import TEST_DATA_PATH
from Utils.LogUtil import log

'''
读取excel中的参数名称和参数值
'''


def get_params(filename, sheet_name, row):
    filePath = TEST_DATA_PATH + f'/{filename}'
    try:
        workbook = openpyxl.load_workbook(filePath)
        sheet = workbook[sheet_name]
        if workbook and sheet:
            columns = sheet.max_column
            params = []
            for c in range(1, columns+1):
                params.append(sheet.cell(row, c).value)
            return params
    except Exception as e:
        log.exception(message='load excel error: {}'.format(e) + 'filePath: {}'.format(filename))


def get_params_value(filename, sheet_name):
    filePath = TEST_DATA_PATH + f'/{filename}'
    try:
        workwook = openpyxl.load_workbook(filePath)
        sheet = workwook[sheet_name]
        rows = sheet.max_row
        columns = sheet.max_column
        column_data = []
        for r in range(2, rows+1):
            row_data = []
            for c in range(1, columns+1):
                row_data.append(sheet.cell(r, c).value)
            column_data.append(row_data)
        return column_data
    except Exception as e:
        log.exception(message='load excel error: {}'.format(e) + 'filePath: {}'.format(filename))


if __name__ == '__main__':
    print(get_params('data.xlsx', 'flow', 1))
    print(get_params_value('data.xlsx', 'flow'))